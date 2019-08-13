import pandas as pd
import openpyxl
# !------------------------------------ Functions -----------------------------------------------#
## * ----- unique vendor list 
def GetUniqueVendor(Ven_list):
    dr = df[df.Location.isin(Ven_list)]
    Ven_list = list(dr['Vendor_Name'].unique())
    print(dr)
    print(Ven_list)
    return (Ven_list)
## ------------------------------------------------
## *--- create work sheet in excelsheets according to vendor name
def Addworksheetxls(List_to_be_added,File_path):
    count = 1
    wb = openpyxl.load_workbook(File_path)
    for G in List_to_be_added:
        print (wb.sheetnames)
        Sheet1 = wb.worksheets[0]
        wb.worksheets[1] = wb.copy_worksheet(Sheet1)
        print (wb.sheetnames)
        wb_sheet = wb.worksheets[count]
        wb_sheet.title = G
        print (wb.sheetnames)
        count += 1
    wb.save(File_path)
    wb.close()    
#---------------------------------------------------------
def Invoice_no(Location):
    ws = openpyxl.load_workbook(filepath_Invoice_number)
    



# !-------------------------------------------------------------------------------------------------#
#----- file path---#
filepath_Invoice_number = r"C:\Users\pv.abhilash\Desktop\vinutha\Import of services invoice series.xlsx"
filepath_Master_excel = r"C:\Users\pv.abhilash\Desktop\vinutha\Att4702tmp.xlsx"
filepath_Office_address = r'C:\Users\pv.abhilash\Desktop\vinutha\Office Addresses with GSTN.xlsx'
filepath_Maharashtra = r"C:\Users\pv.abhilash\Desktop\vinutha\Invoice\Invoices for import of service Maharashtra.xlsx"
filepath_Karnatak = r'C:\Users\pv.abhilash\Desktop\vinutha\Invoice\Invoices for import of service Karnataka.xlsx'
filepath_Gurgoan =  r'C:\Users\pv.abhilash\Desktop\vinutha\Invoice\Invoices for import of service Gurgaon.xlsx'
filepath_Telgana =  r'C:\Users\pv.abhilash\Desktop\vinutha\Invoice\Invoices for import of service Telangana.xlsx'
filepath_invoice_template =  r"C:\Users\pv.abhilash\Desktop\vinutha\Invoice_Format_for_Import_Of_goods.xlsx"
#-------------------#
# ------List--------#
Location_MH_list = ['Maharashtra','Pune SEZ','Mumbai SEZ']
Location_BA_List = ['Bangalore']
Location_Gur_List = ['Gurgaon']
Location_Hyd_list = ['Hyderabad']
#-------------------#
#------Variables----#
typelist = list()
vendorlist = list()
Ven_list = list()
Tel_Loc_inv  = r"RC/TG/"
Mah_Loc_inv  = r"RC/MH/"
Gurg_Loc_inv = r"RC/GG/"
Kar_Loc_inv  = r"RC/KA/" 
invoice_start_Tel = 0
invoice_start_Mah  = 0
invoice_start_Gurg = 0
invoice_start_kar = 0
#-------------------#
# === opening excel =====#
wd = openpyxl.load_workbook(filepath_Office_address)
wd1 = openpyxl.load_workbook(filepath_invoice_template) 
input_file_office = pd.read_excel(filepath_Office_address)    
input_file = pd.read_excel (filepath_Master_excel,"Sheet1") 
#=========================#
#===== Creating Excelfile depnding in location ==== #
loclist = list(input_file['Location'].unique())
print(loclist) 
for i in loclist:
    if i =='Bangalore':
        wd1.save(filepath_Karnatak)
        print(i)
    elif i =='Gurgaon':
        print(i)
        wd1.save(filepath_Gurgoan)
    elif i =='Maharashtra' or i =='Pune SEZ' or i =='Mumbai SEZ' :
        print(i)
        wd1.save(filepath_Maharashtra)    
    elif i =='Hyderabad':
        print(i)
        wd1.save(filepath_Telgana)        
    else:
        print('Please define the new location')
# ============================#
# ===== unique typelist=======#
typelist = list(input_file['Type'].unique())
print(typelist) 
#=============================#
#====== unique Total Vendor list ======#
vendorlist = list(input_file['Vendor_Name'].unique())
print(vendorlist)
#======================================#
# creating dataframe for required columns #
df = pd.DataFrame(input_file,columns = ['Type','Location','Vendor_Name','Net Amount - Local','IGST @ 18%']) 
print(df)
#=======================================#
#==== function calling for creating invoice worksheets======#
GetUniqueVendor(Location_Hyd_list)
Ven_list = GetUniqueVendor(Location_Hyd_list)        #---------- hyderbad
Addworksheetxls(Ven_list,filepath_Telgana)
#-----------
GetUniqueVendor(Location_BA_List)
Ven_list = GetUniqueVendor(Location_BA_List)         #---------- Bangalore
Addworksheetxls(Ven_list,filepath_Karnatak)
#------------
GetUniqueVendor(Location_MH_list)
Ven_list = GetUniqueVendor(Location_MH_list)         #---------- Maharahstra
Addworksheetxls(Ven_list,filepath_Maharashtra)
#-------------
GetUniqueVendor(Location_Gur_List)
Ven_list = GetUniqueVendor(Location_Gur_List)         #--------- Gurgaon
Addworksheetxls(Ven_list,filepath_Gurgoan)
print('Completed')
#====================================#












# for i in loclist:
#         for j in vendorlist:
#             #dr = df.loc[(df.Location== i)&(df.Vendor_Name== j)]
#                 for h in typelist:
#                     dr = df.loc[(df.Location== i)&(df.Vendor_Name== j)&(df.Type== h)]     
#                     amountsum = dr['Net Amount - Local'].sum()
                    
                    # for ind,row in dr.iterrows():
                    #     mount = dr["Net Amount - Local"]
                    #     print (Amount)
                    #     Type = dr['Type']
                    #     Location = dr['Location']
                    #     Vendor = dr['Vendor_Name']
                    #     IGST = dr['IGST @ 18%']

                    #     if Location =='Bangalore':
                    #     wd2=openpyxl.load_workbook(r'C:\Users\pv.abhilash\Desktop\vinutha\Invoice\Invoices for import of service Karnataka.xlsx')
                    #     wd2.create_sheet(title = Vendor)
                    #     print(i)
                    #     elif Location =='Gurgaon':
                    #     print(i)
                    #     wd2=openpyxl.load_workbook(r'C:\Users\pv.abhilash\Desktop\vinutha\Invoice\Invoices for import of service Gurgaon.xlsx')
                    #     wd2.create_sheet(title = Vendor)
                    #     elif Location =='Maharashtra' or i =='Pune SEZ' or i =='Mumbai SEZ' :
                    #     print(i)
                    #     wd2=openpyxl.load_workbook(r'C:\Users\pv.abhilash\Desktop\vinutha\Invoice\Invoices for import of service Maharashtra1.xlsx')    
                    #     wd2.create_sheet(title = Vendor)
                    #     elif Location =='Hyderabad':
                    #     print(i)
                    #     wd2=openpyxl.load_workbook(r'C:\Users\pv.abhilash\Desktop\vinutha\Invoice\Invoices for import of service Telangana.xlsx')        
                    #     wd2.create_sheet(title = Vendor)
                    #     else:
                    #     print('Please define the new location') 
                
               

                    
    
    
